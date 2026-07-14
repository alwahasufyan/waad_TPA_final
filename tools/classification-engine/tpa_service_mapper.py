#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
  TPA Service Mapper  |  مُرتِّب ومُطابِق قوائم الأسعار لنظام التأمين الطبي
=============================================================================
الوظيفة:
  يأخذ ملف Excel "مرجعي" (كتالوج الخدمات الرسمي) + أي قائمة أسعار "فوضوية"
  مُرسَلة من مستشفى/صيدلية/مختبر، ثم يُنتج ملف Excel مُرتَّب وجاهز للاستيراد
  داخل نظام TPA، بنفس أسماء أعمدة المرجع، مع تعبئة:
      - service_code      (الكود)
      - main_category     (التصنيف الرئيسي: إيواء / عيادات خارجية)
      - sub_category      (التصنيف الطبي الرسمي: CAT-*)
      - confidence_score  (درجة الثقة في المطابقة)

  الخدمات التي يكون تصنيفها تقديريًا (غير موثق من عقد/أودو/نوع المرفق) تُحمَّل
  بلون أحمر في نهاية ورقة "الخدمات Services" الموحدة لمراجعة تصنيفها بشريًا
  عبر القوائم المنسدلة المعتمدة.

طرق المطابقة (بالترتيب):
  1) Exact match            - تطابق تام بعد التنظيف
  2) Synonyms               - عبر قاموس المرادفات الطبية (قابل للتوسعة)
  3) Arabic/English Normalize - توحيد الكتابة العربية والإنجليزية
  4) Fuzzy Matching         - تطابق تقريبي (rapidfuzz إن وُجد، وإلا difflib)

طريقة التشغيل:
  python tpa_service_mapper.py \
      --reference  Price_List_Contract_1_Output.xlsx \
      --input      new_hospital_list.xlsx \
      --output     result.xlsx \
      --synonyms   medical_synonyms.json \
      --threshold  85

  (المعاملات --synonyms و --threshold اختيارية)
=============================================================================
"""

import argparse
import json
import os
import re
import sys
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import ingest  # محرك الابتلاع متعدد الصيغ (Excel/PDF/PowerPoint/CSV)

# ---------------------------------------------------------------------------
# محرك المطابقة التقريبية: يفضّل rapidfuzz، ويسقط تلقائيًا إلى difflib
# ---------------------------------------------------------------------------
try:
    from rapidfuzz import fuzz
    from rapidfuzz import process as rf_process    # الأفضل أداءً ودقة
    _FUZZ_ENGINE = "rapidfuzz"
except ImportError:                                # خطة بديلة بدون أي تثبيت
    _FUZZ_ENGINE = "index"


# ===========================================================================
# 1) إعدادات افتراضية  (Defaults)
# ===========================================================================
DEFAULT_THRESHOLD = 85          # أقل درجة ثقة لقبول المطابقة تلقائيًا (0-100)
EXACT_SCORE = 100               # درجة التطابق التام

# الأعمدة المعتمدة في المخرجات (نفس أسماء المرجع تمامًا)
COL_NAME  = "service_name / اسم الخدمة ★"
COL_CODE  = "service_code / الكود"
COL_PRICE = "contract_price / سعر العقد"
COL_MAIN  = "main_category / التصنيف الرئيسي"
COL_SUB   = "sub_category / البند (التصنيف الفرعي)"
COL_NOTE  = "notes / ملاحظات"
STD_COLS  = [COL_NAME, COL_CODE, COL_PRICE, COL_MAIN, COL_SUB, COL_NOTE]

# أعمدة مساعدة للمراجع (تأتي بعد أعمدة الاعتماد الستة)
COL_REASON  = "سبب النتيجة / Reason"
COL_STATUS  = "حالة التصنيف / Status"
STATUS_OK     = "✔ موثق"
STATUS_REVIEW = "⚠ راجِع التصنيف"
COL_CONF   = "confidence_score / درجة الثقة"
COL_METHOD = "match_method / طريقة المطابقة"
COL_REF    = "المرجع أو أقرب اقتراح / Reference"
COL_SRC    = "source / المصدر"
ALL_COLS = STD_COLS + [COL_STATUS, COL_REASON, COL_CONF, COL_METHOD,
                       COL_REF, COL_SRC]



# ===========================================================================
# 2) التنظيف والتوحيد  (Normalization)
# ===========================================================================
_AR_DIACRITICS = re.compile(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]")
_TATWEEL = re.compile(r"\u0640")
_NON_KEEP = re.compile(r"[^0-9a-z\u0621-\u064A\s]")   # احتفظ: أرقام/لاتيني/عربي/مسافة
_MULTISPACE = re.compile(r"\s+")


def clean_cell(v) -> str:
    """يحوّل أي قيمة خلية إلى نص نظيف، ويعامل الفراغ/nan/None كنص فارغ."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s


# كلمات وصف عامة تُزال قبل المطابقة (لا تحمل معنى تمييزيًا للخدمة)
_STOPWORDS = {"serum", "plasma", "plain", "playn", "sample", "sampel",
              "specimen", "test", "in", "by", "for", "the", "of",
              "عينه", "عينة", "في", "تحليل"}


def normalize(text) -> str:
    """تنظيف وتوحيد النص (عربي + إنجليزي) لجعله قابلًا للمقارنة."""
    if text is None:
        return ""
    s = str(text)
    s = unicodedata.normalize("NFKC", s)      # توحيد الأشكال (أرقام/حروف)
    s = s.lower()                             # توحيد اللاتيني
    s = _AR_DIACRITICS.sub("", s)             # إزالة التشكيل
    s = _TATWEEL.sub("", s)                   # إزالة التطويل ـــ
    # توحيد الألف والهمزات والياء والتاء المربوطة
    s = re.sub(r"[إأآٱ]", "ا", s)
    s = s.replace("ى", "ي").replace("ئ", "ي").replace("ؤ", "و")
    s = s.replace("ة", "ه")
    s = _NON_KEEP.sub(" ", s)                 # إزالة الرموز والترقيم
    s = _MULTISPACE.sub(" ", s).strip()       # دمج المسافات
    # أزل كلمات الوصف العامة (مع الإبقاء على كلمة واحدة على الأقل)
    toks = [t for t in s.split() if t not in _STOPWORDS]
    return " ".join(toks) if toks else s


# ===========================================================================
# 3) قاموس المرادفات  (Synonyms)
# ===========================================================================
def load_synonyms(path):
    """يحمّل قاموس المرادفات ويبني خريطة (مرادف مُنظَّف -> الصيغة المعتمدة المُنظَّفة)."""
    mapping = {}
    if not path or not os.path.exists(path):
        return mapping
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    for canonical, variants in raw.items():
        if canonical.startswith("_"):         # تجاهل مفاتيح الوصف (_description …)
            continue
        canon_n = normalize(canonical)
        if not canon_n:
            continue
        mapping[canon_n] = canon_n
        for v in variants:
            vn = normalize(v)
            if vn:
                mapping[vn] = canon_n
    return mapping


def apply_synonyms(norm_text, syn_map):
    """يستبدل أي كلمة/عبارة لها مرادف معروف بصيغتها المعتمدة (مطابقة كلمات كاملة)."""
    if not syn_map or not norm_text:
        return norm_text
    words = norm_text.split()
    out = []
    i = 0
    n = len(words)
    while i < n:
        replaced = False
        # جرّب عبارات من 3 كلمات ثم 2 ثم 1 (الأطول أولًا)
        for span in (3, 2, 1):
            if i + span <= n:
                phrase = " ".join(words[i:i + span])
                if phrase in syn_map:
                    out.append(syn_map[phrase])
                    i += span
                    replaced = True
                    break
        if not replaced:
            out.append(words[i])
            i += 1
    return _MULTISPACE.sub(" ", " ".join(out)).strip()


# ===========================================================================
# 4) تحميل المرجع وبناء الفهرس  (Reference index)
# ===========================================================================
def build_reference(ref_path, syn_map):
    """يقرأ المرجع ويبني فهرسًا: لكل اسم مُنظَّف -> (code, main, sub, الاسم الأصلي)."""
    df = pd.read_excel(ref_path, dtype=str)
    # تطابق الأعمدة بالاسم؛ وإن اختلفت قليلًا نأخذ أول 6 أعمدة بالترتيب
    cols = list(df.columns)
    if not all(c in cols for c in STD_COLS):
        if len(cols) >= 5:
            rename = dict(zip(cols[:6], STD_COLS[:len(cols[:6])]))
            df = df.rename(columns=rename)
        else:
            raise ValueError("ملف المرجع لا يحتوي على الأعمدة المتوقعة.")

    exact = {}      # اسم مُنظَّف -> سجل
    exact_syn = {}  # اسم مُنظَّف + مرادفات -> سجل
    choices = {}    # اسم مُنظَّف -> الاسم المُنظَّف (لقائمة المطابقة التقريبية)

    for _, row in df.iterrows():
        name = row.get(COL_NAME)
        if name is None or str(name).strip() == "" or str(name) == "nan":
            continue
        rec = {
            "name": clean_cell(name),
            "code": clean_cell(row.get(COL_CODE)),
            "main": clean_cell(row.get(COL_MAIN)),
            "sub":  clean_cell(row.get(COL_SUB)),
        }
        nn = normalize(name)
        if not nn:
            continue
        # أول ظهور يفوز (المرجع قد يكرر نفس الاسم بأكواد مختلفة)
        exact.setdefault(nn, rec)
        choices.setdefault(nn, nn)
        nsyn = apply_synonyms(nn, syn_map)
        exact_syn.setdefault(nsyn, rec)
        choices.setdefault(nsyn, nsyn)

    choice_list = list(choices.keys())
    return df, exact, exact_syn, choices, choice_list


# ===========================================================================
# 4.4) التصنيفات المعتمدة  (Approved categories) ⭐
# ===========================================================================
# ملف official_taxonomy.json المولد من وثيقة TAX-1 هو المصدر الوحيد للتصنيفات.
# كل خدمة في المخرجات يجب أن تُسند إلى أحد هذه التصنيفات حصرًا.
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CATEGORIES_FILE = os.path.join(_BASE_DIR, "official_taxonomy.json")


def load_approved_categories(path):
    """Loads the immutable TAX-1 CAT catalogue and rejects every other format."""
    if not str(path).lower().endswith(".json"):
        raise ValueError("TAX-1 accepts official_taxonomy.json only")
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    cats = {}
    for row in payload.get("medical_categories", []):
        code = clean_cell(row.get("code"))
        name = clean_cell(row.get("name_ar"))
        if not re.fullmatch(r"CAT-[A-Z0-9-]+", code) or not name:
            raise ValueError(f"Non-official medical category in TAX-1 source: {code!r}")
        cats[code] = {
            "name": name,
            "main": "",
            "contexts": list(row.get("contexts") or []),
            "label": f"{code} - {name}",
        }
    if len(cats) != 33:
        raise ValueError(f"Official taxonomy must contain exactly 33 CAT entries: {path}")
    return cats


# قاعدة معرفة نظام أودو السابق: اسم مُطبَّع -> تصنيف معتمد موثق
# (تُبنى بـ build_odoo_kb.py من تصديري product.product / product.template)
DEFAULT_KB_FILE = os.path.join(_BASE_DIR, "official_knowledge.json")


def load_knowledge_base(path, allowed_codes=None):
    """Load the operational knowledge file and fail on non-official codes."""
    if not path or not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        knowledge = json.load(f)
    if allowed_codes is not None:
        invalid = sorted({row.get("cat") for row in knowledge.values()}
                         - set(allowed_codes))
        if invalid:
            raise ValueError(f"Non-official categories in knowledge file: {invalid}")
    return knowledge


# ===========================================================================
# 4.5) المُصنِّف الطبي  (Medical category classifier)
# ===========================================================================
# TAX-1 official rules. Order is safety-critical: specific
# dental/psychiatric/maternity categories precede their broader neighbours.
_CATEGORY_RULES = [
    ("CAT-DENT-EMERG", ["طوارئ اسنان", "اسنان طوارئ", "dental emergency"]),
    ("CAT-DENT-IMPLANT", ["زراعه سن", "زرعه سن", "dental implant", "implant"]),
    ("CAT-DENT-ORTHO", ["تقويم اسنان", "orthodont"]),
    ("CAT-DENT-PROSTHO", ["تركيب اسنان", "طقم اسنان", "تلبيس", "تاج", "جسر", "denture", "crown", "bridge", "prosth"]),
    ("CAT-DENT-ROUTINE", ["حشو", "خلع", "تنظيف اسنان", "علاج عصب", "root canal", "filling", "extraction", "scaling", "dental"]),
    ("CAT-PSYCH-DRUG", ["دواء نفسي", "ادويه نفسيه", "antidepress", "antipsychotic", "psychiatric drug"]),
    ("CAT-PSYCH-SESS", ["جلسه نفسيه", "جلسات نفسيه", "psychotherapy", "counselling", "counseling"]),
    ("CAT-MAT-CS", ["ولاده قيصريه", "قيصري", "cesarean", "caesarean"]),
    ("CAT-MAT-NORMAL", ["ولاده طبيعيه", "normal delivery", "nvd"]),
    ("CAT-MAT-COMP", ["مضاعفات حمل", "مضاعفات ولاده", "placenta", "ectopic", "حمل خارج الرحم"]),
    ("CAT-ICU", ["عنايه فايقه", "عنايه مركزه", "intensive care", "icu", "nicu"]),
    ("CAT-CCU", ["عنايه القلب", "cardiac care unit", "ccu"]),
    ("CAT-IMG-ADV", ["رنين مغناطيسي", "تصوير مقطعي", "اشعه مقطعيه", "mri", "ct scan", "computed tomography"]),
    ("CAT-IMG-DIAG", ["اشعه سينيه", "اشعه تشخيصيه", "x ray", "x-ray", "ultrasound", "سونار", "doppler"]),
    ("CAT-LAB", ["تحليل", "مختبر", "cbc", "pcr", "culture", "blood test", "urine", "stool", "hormone"]),
    ("CAT-ANESTHESIA", ["تخدير", "بنج", "anesth", "sedation"]),
    ("CAT-SURG-MAT", ["مواد جراحيه", "معدات جراحيه", "surgical material"]),
    ("CAT-SURGERY", ["عمليه جراحيه", "جراحه", "استئصال", "surgery", "surgical", "ectomy", "otomy"]),
    ("CAT-PRACT-FEE", ["رسوم طبيب", "رسوم اخصائي", "رسوم استشاري", "consultation fee", "doctor fee"]),
    ("CAT-DIAGNOSTIC", ["كشف تشخيصي", "فحص تشخيصي", "diagnostic examination"]),
    ("CAT-DAY-CARE", ["رعايه يوميه", "علاج يومي", "day care"]),
    ("CAT-ROOM", ["غرفه خاصه", "اقامه", "ايواء", "سرير", "private room"]),
    ("CAT-AMBULANCE", ["اسعاف", "ambulance"]),
    ("CAT-HOME-NURSING", ["تمريض منزلي", "تمريض في المنزل", "home nursing"]),
    ("CAT-PHYSIO", ["علاج طبيعي", "تاهيل", "physiother", "rehabilit"]),
    ("CAT-TRANSPLANT", ["زرع عضو", "زراعه عضو", "organ transplant"]),
    ("CAT-ONCOLOGY", ["اورام", "كيماوي", "oncolog", "chemother", "radiother"]),
    ("CAT-DIALYSIS", ["غسيل كلوي", "غسيل الكلى", "dialysis"]),
    ("CAT-DME", ["كرسي متحرك", "جهاز تعويضي", "durable medical equipment", "dme"]),
    ("CAT-MED-SUP", ["مستلزمات طبيه", "medical supplies", "consumables"]),
    ("CAT-DRUG", ["دواء", "ادويه", "صيدليه", "medication", "pharmacy"]),
    ("CAT-OPT", ["نظاره", "نظارات", "عدسات لاصقه", "spectacle", "eyeglass", "contact lens"]),
    ("CAT-EYE-EXAM", ["كشف عيون", "فحص عيون", "eye examination", "ophthalmic exam"]),
]


class CategoryClassifier:
    """يصنّف أي خدمة بقواعد طبية على التصنيفات **المعتمدة** حصرًا."""

    def __init__(self, cats):
        self.cats = cats         # {كود معتمد: {name, main, label}}
        # TAX-1: no GENERAL/OUTPATIENT/default medical category is permitted.
        self.default_key = None
        # طبِّع الكلمات المفتاحية مرة واحدة
        self.rules = [(k, [normalize(w) for w in kws if normalize(w)])
                      for k, kws in _CATEGORY_RULES if k in cats]

    @staticmethod
    def _stem(tok):
        """يجرّد أدوات التعريف من كلمة عربية (الـ/وال/بال...) لتسهيل المطابقة."""
        for pre in ("وال", "فال", "بال", "كال", "لل", "ال"):
            if tok.startswith(pre) and len(tok) > len(pre):
                return tok[len(pre):]
        if len(tok) > 1 and tok[0] in "وفبكل":
            return tok[1:]
        return tok

    def _ar_token_match(self, kw, stems):
        return kw in stems

    def _hit(self, kw, text, tokens, stems):
        if " " in kw:                       # عبارة
            if kw.isascii():
                return kw in text           # لاتينية: مطابقة جزئية
            # عربية: كل كلماتها موجودة كجذور أو ككلمات كاملة
            return all(self._stem(w) in stems or w in tokens
                       for w in kw.split())
        if kw.isascii():
            if len(kw) >= 5:                # لاحقة/مقطع (ectomy, scope) → جزئي
                return kw in text
            return kw in tokens             # كلمة قصيرة (ct, mri) → كاملة
        # عربي مفرد: كجذر أو ككلمة كاملة (يحمي كلمات تبدأ بـ و/ف/ب مثل «وتد»)
        return kw in stems or kw in tokens

    def _cat(self, key):
        c = self.cats[key]
        return c["main"], c["label"], key

    def classify_rules(self, raw_text):
        """يطبّق القواعد الطبية فقط؛ يعيد (main, sub, catkey) أو None."""
        t = normalize(raw_text)
        if not t:
            return None
        toks = set(t.split())
        stems = {self._stem(x) for x in toks}
        for key, kws in self.rules:
            for kw in kws:
                if self._hit(kw, t, toks, stems):
                    return self._cat(key)
        return None

    def classify(self, raw_text, default_key=None):
        """يعيد (main, sub, catkey) — قواعد طبية ثم التصنيف الافتراضي."""
        hit = self.classify_rules(raw_text)
        if hit:
            return hit
        dk = default_key if (default_key in self.cats) else None
        return self._cat(dk) if dk else ("", "", None)

    def to_approved(self, sub_raw, service_text):
        """Accept an official CAT-* reference or reclassify its service text."""
        m = re.match(r"\s*(CAT-[A-Z0-9-]+)", clean_cell(sub_raw))
        if m:
            code = m.group(1)
            if code in self.cats:
                c = self.cats[code]
                return c["main"], c["label"], ""
        hit = self.classify_rules(service_text)
        if hit:
            main, sub, _ = hit
            return main, sub, "يتطلب تأكيد المراجع: التصنيف استُنتج من قاعدة نصية"
        main, sub, _ = ("", "", None)
        return main, sub, ""


# ===========================================================================
# 5) المطابقة التقريبية  (Fuzzy)
# ===========================================================================
import difflib
from collections import defaultdict


class FuzzyIndex:
    """فهرس كلمات مقلوب لمطابقة تقريبية سريعة بدون أي مكتبات خارجية.

    يحجب المقارنات إلى المرشّحين الذين يشاركون كلمة واحدة على الأقل، ثم
    يرتّبهم بمعامل تشابه المجموعات (Jaccard) ويصقل الأفضل بـ difflib.
    """

    def __init__(self, choice_list):
        self.choices = choice_list
        self.tokens = [set(c.split()) for c in choice_list]
        self.inverted = defaultdict(list)
        for i, toks in enumerate(self.tokens):
            for t in toks:
                self.inverted[t].append(i)

    def best(self, query_norm, topk=5):
        if not query_norm:
            return None, 0, {}
        q = set(query_norm.split())
        if not q:
            return None, 0, {}
        # اجمع المرشّحين الذين يشاركون كلمة على الأقل
        cand = set()
        for t in q:
            cand.update(self.inverted.get(t, ()))
        if not cand:
            return None, 0, {}
        # رتّب بـ Jaccard ثم اصقل أفضل القلّة بـ difflib
        scored = []
        for i in cand:
            inter = len(q & self.tokens[i])
            union = len(q | self.tokens[i])
            scored.append((inter / union if union else 0, i))
        scored.sort(reverse=True)
        best_score, best_idx, best_comp = 0, None, {}
        for jac, i in scored[:topk]:
            seq = difflib.SequenceMatcher(None, query_norm, self.choices[i]).ratio()
            inter = len(q & self.tokens[i])
            # الاحتواء الجزئي — يُطبَّق فقط عندما يكون لكلا الطرفين كلمتان فأكثر
            # (يمنع مطابقة جملة طويلة بمدخل مرجعي مفرد عام مثل "IgM")
            if len(q) >= 2 and len(self.tokens[i]) >= 2:
                contain = inter / min(len(q), len(self.tokens[i]))
            else:
                contain = 0
            sc = max(jac, seq, contain * 0.9) * 100
            if sc > best_score:
                best_score, best_idx = sc, i
                best_comp = {"jac": round(jac * 100), "seq": round(seq * 100),
                             "contain": round(contain * 100)}
        return (self.choices[best_idx] if best_idx is not None else None,
                round(best_score, 1), best_comp)


def fuzzy_best(query_norm, fidx, ref_lookup):
    """يعيد (السجل, الدرجة, الاسم_المطابق, مكوّنات الدرجة)."""
    if not query_norm:
        return None, 0, None, {}
    if _FUZZ_ENGINE == "rapidfuzz":
        best = rf_process.extractOne(query_norm, fidx.choices,
                                  scorer=fuzz.token_set_ratio)
        if best is None:
            return None, 0, None, {}
        matched_norm, score = best[0], round(best[1], 1)
        comp = {"jac": "", "seq": "", "contain": ""}
    else:
        matched_norm, score, comp = fidx.best(query_norm)
        if matched_norm is None:
            return None, 0, None, {}
    return ref_lookup.get(matched_norm), score, matched_norm, comp


def _signal_text(comp):
    """يصف أي إشارة رفعت الدرجة (كلمات/حروف/احتواء)."""
    if not comp or comp.get("jac") == "":
        return "تشابه نصّي"
    jac, seq, con = comp.get("jac", 0), comp.get("seq", 0), comp.get("contain", 0)
    top = max(jac, seq, con * 0.9)
    if con * 0.9 == top and con >= 60:
        return "كلمات الخدمة جزء من اسم بالمرجع (احتواء)"
    if jac >= seq:
        return f"تطابق في الكلمات ({jac}%)"
    return f"تشابه في الحروف ({seq}%) — قد تختلف كلمة، تحقّق أنها لا تغيّر المعنى"


def build_reason(res, threshold, is_review):
    """يبني عبارة «سبب النتيجة بنظرة» للمراجع."""
    method, score, comp = res["method"], res["score"], res.get("comp", {})
    if "تطابق تام" in method:
        return "✔ تطابق تام مع المرجع"
    if "مرادف" in method:
        return "✔ مطابقة عبر مرادف معروف"
    # تقريبي
    if is_review:
        if score == 0 or not res["matched"]:
            return "✖ لا يوجد ما يقابلها في المرجع — غالبًا غير متعاقد عليها"
        if score < 50:
            return f"بعيدة عن كل خدمات المرجع ({score}%) — راجِع: قد تكون غير موجودة"
        if score < 70:
            return (f"صياغة مختلفة/كلمات زائدة ({score}%) — قد تكون صحيحة، "
                    f"أكّد الاقتراح أو عدّله")
        return f"قريبة جدًا من الاقتراح ({score}%) — راجِع وأكّد يدويًا"
    # مقبولة في النتيجة
    sig = _signal_text(comp)
    if score >= 90:
        return f"موثوق ({score}%): {sig}"
    return f"⚠ ثقة حدّية ({score}%): {sig} — يُستحسن نظرة سريعة"


def _ig_class(text):
    """يستخرج فئة الجسم المضاد (igg/igm/iga/ige) إن وُجدت."""
    found = set(re.findall(r"\big[gma e]\b|\bigg\b|\bigm\b|\biga\b|\bige\b", text))
    cls = set()
    for f in found:
        f = f.replace(" ", "")
        if f in ("igg", "igm", "iga", "ige"):
            cls.add(f)
    return cls


def match_service(raw_name, syn_map, exact, exact_syn, fidx, ref_lookup):
    """يطابق خدمة واحدة ويعيد قاموس النتيجة."""
    nn = normalize(raw_name)
    nsyn = apply_synonyms(nn, syn_map)

    # 1) تطابق تام
    if nn in exact:
        return _result(exact[nn], EXACT_SCORE, "exact / تطابق تام",
                       exact[nn]["name"], {})
    # 2) تطابق تام بعد المرادفات
    if nsyn in exact_syn:
        return _result(exact_syn[nsyn], EXACT_SCORE, "synonym / مرادف",
                       exact_syn[nsyn]["name"], {})
    # 3+4) تقريبي (نجرّب النص العادي والنص بعد المرادفات ونأخذ الأعلى)
    rec1, sc1, m1, c1 = fuzzy_best(nn,   fidx, ref_lookup)
    rec2, sc2, m2, c2 = fuzzy_best(nsyn, fidx, ref_lookup)
    if sc2 > sc1:
        rec, sc, matched, comp = rec2, sc2, m2, c2
    else:
        rec, sc, matched, comp = rec1, sc1, m1, c1

    # حارس فئات الأجسام المضادة: لا تَقبل خلط IgG/IgM/IgA/IgE
    if rec is not None:
        q_cls = _ig_class(nn) or _ig_class(nsyn)
        r_cls = _ig_class(matched or "")
        if q_cls and r_cls and not (q_cls & r_cls):
            sc = min(sc, 60)   # اخفض الدرجة لتذهب للمراجعة البشرية

    return _result(rec, sc, "fuzzy / تقريبي",
                   rec["name"] if rec else "", comp)


def _result(rec, score, method, matched_name, comp):
    if rec is None:
        return {"code": "", "main": "", "sub": "", "score": 0,
                "method": "none / لا يوجد", "matched": "", "comp": {}}
    return {"code": rec["code"], "main": rec["main"], "sub": rec["sub"],
            "score": score, "method": method, "matched": matched_name,
            "comp": comp}


# ===========================================================================
# 7) كتابة المخرجات  (Excel output with formatting)
# ===========================================================================
from openpyxl.worksheet.datavalidation import DataValidation

HDR_FILL   = PatternFill("solid", fgColor="1F4E78")   # أزرق غامق للعناوين
IMPORT_FILL = PatternFill("solid", fgColor="1F6E43")  # أخضر = الأعمدة الست للاعتماد
HELPER_FILL = PatternFill("solid", fgColor="7F7F7F")  # رمادي = أعمدة مساعدة
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HIGH_FILL  = PatternFill("solid", fgColor="C6EFCE")   # أخضر = ثقة عالية
MID_FILL   = PatternFill("solid", fgColor="FFEB9C")   # أصفر = ثقة متوسطة
LOW_FILL   = PatternFill("solid", fgColor="FFC7CE")   # أحمر = ثقة منخفضة
BODY_FONT  = Font(name="Arial", size=10)


def _style_sheet(ws, n_cols, conf_col=None, import_cols=0):
    """import_cols: عدد الأعمدة الأولى التي تطابق ملف الاعتماد (تُلوَّن مميَّزة)."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = IMPORT_FILL if (import_cols and c <= import_cols) else \
            (HELPER_FILL if import_cols else HDR_FILL)
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        maxlen = max([len(str(ws.cell(row=r, column=c).value or ""))
                      for r in range(1, min(ws.max_row, 200) + 1)] + [10])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 12), 55)
    for r in range(2, ws.max_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
        if conf_col:
            cell = ws.cell(row=r, column=conf_col)
            try:
                v = float(cell.value)
                cell.fill = HIGH_FILL if v >= 90 else MID_FILL if v >= 75 else LOW_FILL
                cell.alignment = Alignment(horizontal="center")
            except (TypeError, ValueError):
                pass


def _add_instructions(wb):
    ws = wb.create_sheet("تعليمات الاستخدام")
    ws.sheet_view.rightToLeft = True
    lines = [
        ("كيفية استخدام هذا الملف", True),
        ("", False),
        ("ورقة «الخدمات» واحدة تجمع كل شيء: الأعمدة الستة الأولى (بالأخضر) "
         "مطابقة تمامًا لأعمدة ملف الرفع — انسخها كما هي بعد إنهاء المراجعة.", False),
        ("", False),
        ("الصفوف البيضاء: تصنيفها موثق (مرجع العقد / نظام أودو السابق / نوع "
         "المرفق) — لا تحتاج شيئًا.", False),
        ("الصفوف الحمراء (في نهاية الجدول): تصنيفها تقديري — هذه فقط مهمة "
         "المراجع: تأكد من التصنيف الرئيسي والفرعي وصحّحهما من القائمة "
         "المنسدلة في الخليتين إن لزم.", False),
        ("", False),
        ("عمودا التصنيف الرئيسي (D) والفرعي (E): كل خلية فيها قائمة منسدلة "
         "بالتصنيفات المعتمدة حصرًا — اختر منها ولا تكتب يدويًا.", False),
        ("تذكر: التصنيف الرئيسي يجب أن يطابق أب التصنيف الفرعي "
         "(إيواء أو عيادات خارجية).", False),
        ("عمود «حالة التصنيف» وعمود «سبب النتيجة» يوضحان مصدر كل تصنيف "
         "(مطابقة عقد / أودو / نوع المرفق / تقديري).", False),
        ("", False),
        ("بعد المراجعة: ظلّل الأعمدة الستة الأولى كاملة وانسخها إلى ملف "
         "الرفع — أو شغّل: python collect_approved.py \"اسم_الملف.xlsx\"", False),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = Font(name="Arial", size=13 if bold else 10, bold=bold,
                      color="1F4E78" if bold else "000000")
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 95


DATA_SHEET = "الخدمات Services"
LISTS_SHEET = "قوائم التصنيفات"


def write_output(output_path, df_all, n_review, summary, cats):
    """يكتب ورقة بيانات واحدة: الموثق أولًا ثم صفوف المراجعة محمّرة بالنهاية،
    مع قوائم منسدلة للتصنيف الرئيسي والفرعي من التصنيفات المعتمدة حصرًا."""
    df_all.to_excel(output_path, sheet_name=DATA_SHEET, index=False)
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as xl:
        summary.to_excel(xl, sheet_name="ملخص Summary", index=False)

    wb = load_workbook(output_path)
    ws = wb[DATA_SHEET]
    n_rows = df_all.shape[0]
    n_cols = df_all.shape[1]
    conf_col = list(df_all.columns).index(COL_CONF) + 1
    _style_sheet(ws, n_cols, import_cols=6, conf_col=conf_col)

    # ظلّل صفوف المراجعة (آخر n_review صفًا) بالأحمر
    if n_review > 0:
        first_red = n_rows - n_review + 2      # +2: العناوين صف 1 والبيانات من 2
        for r in range(first_red, n_rows + 2):
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = LOW_FILL

    # ورقة قوائم التصنيفات المعتمدة (مخفية) — مصدر القوائم المنسدلة
    wl = wb.create_sheet(LISTS_SHEET)
    mains = sorted({c["main"] for c in cats.values() if c["main"]})
    subs = [cats[k]["label"] for k in sorted(cats.keys())]
    for i, v in enumerate(mains, 1):
        wl.cell(row=i, column=1, value=v)
    for i, v in enumerate(subs, 1):
        wl.cell(row=i, column=2, value=v)
    wl.sheet_state = "hidden"

    # القوائم المنسدلة على كامل عمودي التصنيف
    if n_rows > 0:
        main_col = get_column_letter(list(df_all.columns).index(COL_MAIN) + 1)
        sub_col = get_column_letter(list(df_all.columns).index(COL_SUB) + 1)
        dv_main = DataValidation(
            type="list", formula1=f"'{LISTS_SHEET}'!$A$1:$A${len(mains)}",
            allow_blank=True, showErrorMessage=True)
        dv_main.error = "اختر من التصنيفات الرئيسية المعتمدة فقط"
        dv_sub = DataValidation(
            type="list", formula1=f"'{LISTS_SHEET}'!$B$1:$B${len(subs)}",
            allow_blank=True, showErrorMessage=True)
        dv_sub.error = "اختر من التصنيفات الفرعية المعتمدة فقط"
        ws.add_data_validation(dv_main)
        ws.add_data_validation(dv_sub)
        dv_main.add(f"{main_col}2:{main_col}{n_rows + 1}")
        dv_sub.add(f"{sub_col}2:{sub_col}{n_rows + 1}")
        # فلتر تلقائي للفرز والترشيح
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"

    _style_sheet(wb["ملخص Summary"], summary.shape[1])
    _add_instructions(wb)
    order = [DATA_SHEET, "ملخص Summary", "تعليمات الاستخدام", LISTS_SHEET]
    wb._sheets.sort(key=lambda s: order.index(s.title)
                    if s.title in order else 99)
    wb.save(output_path)


# ===========================================================================
# 8) المعالجة الرئيسية  (Main pipeline)
# ===========================================================================
# تلميحات نوع المرفق لا تنشئ تصنيفًا افتراضيًا؛ تساعد فقط في تضييق القواعد.
GROUP_HINTS = ("dental", "optics", "physio")
_HINT_DEFAULT_KEY = {}


def process(reference, input_file, output, synonyms_path, threshold,
            code_prefix="NEW", categories_path=None, group_hint=None):
    print(f"• محرك المطابقة التقريبية: {_FUZZ_ENGINE}")
    syn_map = load_synonyms(synonyms_path)
    print(f"• مرادفات محمّلة: {len(syn_map)} مدخلًا")

    # التصنيفات المعتمدة — المصدر الوحيد لتصنيفات المخرجات
    categories_path = categories_path or DEFAULT_CATEGORIES_FILE
    cats = load_approved_categories(categories_path)
    print(f"• التصنيفات المعتمدة: {len(cats)} تصنيفًا من "
          f"{os.path.basename(categories_path)}")

    df_ref, exact, exact_syn, choices, choice_list = build_reference(reference, syn_map)
    ref_lookup = choices  # name_norm -> name_norm ; نحتاج خريطة norm->record
    # ابنِ خريطة norm -> record من exact + exact_syn
    ref_record = {}
    ref_record.update(exact)
    for k, v in exact_syn.items():
        ref_record.setdefault(k, v)
    print(f"• خدمات المرجع: {len(df_ref)} صفًا ({len(exact)} اسمًا فريدًا)")
    fidx = FuzzyIndex(choice_list)   # فهرس ضبابي يُبنى مرة واحدة
    catclf = CategoryClassifier(cats)   # مُصنِّف طبي على التصنيفات المعتمدة
    print(f"• المصنّف الطبي: {len(catclf.rules)} قاعدة "
          f"(افتراضي: {catclf.default_key})")

    df_in = ingest.load_records(input_file)
    srcs = df_in["source"].nunique() if "source" in df_in.columns else 1
    print(f"• تمت قراءة الإدخال ({srcs} مصدرًا/ورقة) — خدمات واردة: {len(df_in)}")

    if group_hint:
        print(f"• تلميح المجموعة: {group_hint}")
    hint_default = _HINT_DEFAULT_KEY.get(group_hint)

    kb = load_knowledge_base(DEFAULT_KB_FILE, set(cats))
    if kb:
        print(f"• قاعدة معرفة أودو: {len(kb)} اسم خدمة مصنف")

    def kb_lookup(*names):
        """يبحث في قاعدة المعرفة الرسمية باسم الخدمة (عربي أو إنجليزي).

        يعيد (main, label, catkey) أو None.
        """
        for cand in names:
            e = kb.get(normalize(clean_cell(cand)))
            if e and e.get("cat") in cats:
                c = cats[e["cat"]]
                return c["main"], c["label"], e["cat"]
        return None

    def final_category(ref_sub, name, name_alt=""):
        """التصنيف المعتمد النهائي بالأولوية:
        تلميح المجموعة ← المرجع الرسمي ← قاعدة المعرفة ← القواعد الطبية.

        يعيد (main, sub_label, note, source) حيث source مصدر التصنيف:
        hint (نوع المرفق) / ref (مرجع العقد) / odoo (النظام السابق)
        / rule (قواعد طبية) — المصادر الثلاثة الأولى
        **موثوقة** والصف معها جاهز للرفع دون مراجعة تصنيف.
        """
        text = f"{name} {name_alt}"
        hit = catclf.classify_rules(text)
        if group_hint == "dental":
            # مراكز الأسنان: كل شيء أسنان — تركيبات إن دلّت القواعد وإلا روتيني
            if hit and hit[2].startswith("CAT-DENT-"):
                return hit[0], hit[1], "", "rule"
            return "", "", "", "unresolved"
        if group_hint == "optics" and hit and hit[2] in ("CAT-OPT", "CAT-EYE-EXAM"):
            return hit[0], hit[1], "", "hint"   # النظارات/الكشوف حسب نوع المرفق
        if clean_cell(ref_sub):           # مطابَقة بالمرجع → حوّل للمعتمد
            official = re.match(r"\s*(CAT-[A-Z0-9-]+)", clean_cell(ref_sub))
            m, s, note = catclf.to_approved(ref_sub, text)
            return m, s, note, ("ref" if official and official.group(1) in cats else "rule")
        found = kb_lookup(name, name_alt)
        if found:
            return found[0], found[1], "التصنيف من قاعدة المعرفة الرسمية", "knowledge"
        if hit:
            return hit[0], hit[1], "", "rule"
        dk = hint_default if (hint_default in cats) else catclf.default_key
        if dk:
            m, s, _ = catclf._cat(dk)
            return m, s, "", ("hint" if dk == hint_default else "default")
        return "", "", "", "unresolved"

    matched_rows, review_rows = [], []
    gen_counter = [0]
    n_cat_trusted = [0]   # صفوف جاهزة بتصنيف موثق (دون مطابقة سعرية بالعقد)

    def resolve_code(*candidates):
        """يعيد أول كود غير فارغ، وإلا يولّد كودًا جديدًا في عمود الكود."""
        for c in candidates:
            c = clean_cell(c)
            if c:
                return c
        gen_counter[0] += 1
        return f"{code_prefix}{gen_counter[0]:05d}"

    for _, row in df_in.iterrows():
        name = clean_cell(row["name"])
        if not name:
            continue
        price = "" if pd.isna(row["price"]) else row["price"]
        source = row.get("source", "")
        src_code = clean_cell(row.get("code", ""))   # كود المستشفى الأصلي
        # جرّب الاسم الأساسي والاسم الثانوي (مثل العربي/الإنجليزي) وخذ الأفضل
        res = match_service(name, syn_map, exact, exact_syn, fidx, ref_record)
        name_alt = clean_cell(row.get("name_alt", ""))
        if name_alt and name_alt.lower() != name.lower():
            res_alt = match_service(name_alt, syn_map, exact, exact_syn,
                                    fidx, ref_record)
            if res_alt["score"] > res["score"]:
                res = res_alt
        # اعرض الاسمين معًا في عمود الاسم إن اختلفا (أوضح للمراجع)
        display_name = name if not name_alt or name_alt.lower() == name.lower() \
            else f"{name} | {name_alt}"
        if res["score"] >= threshold:
            # مطابقة سعرية بالعقد: كود المرجع ← كود المصدر ← كود مولّد
            amain, asub, anote, category_source = final_category(res["sub"], name, name_alt)
            matched_rows.append({
                COL_NAME: display_name,
                COL_CODE: resolve_code(res["code"], src_code),
                COL_PRICE: price,
                COL_MAIN: amain,
                COL_SUB: asub,
                COL_NOTE: anote,
                COL_STATUS: STATUS_OK if category_source in ("ref", "odoo") else STATUS_REVIEW,
                COL_REASON: build_reason(res, threshold, is_review=False),
                COL_CONF: res["score"],
                COL_METHOD: res["method"],
                COL_REF: res["matched"],
                COL_SRC: source,
            })
            continue

        cmain, csub, cnote, csrc = final_category("", name, name_alt)
        if csrc in ("hint", "odoo"):
            # تصنيف موثق (نوع المرفق / النظام السابق) — المعيار المعتمد
            # للجاهزية هو صحة التصنيف، فالصف جاهز للرفع بكوده وسعره الواردين
            n_cat_trusted[0] += 1
            src_label = ("تصنيف موثق حسب نوع المرفق" if csrc == "hint"
                         else "تصنيف موثق من نظام أودو السابق")
            matched_rows.append({
                COL_NAME: display_name,
                COL_CODE: resolve_code(src_code),
                COL_PRICE: price,
                COL_MAIN: cmain,
                COL_SUB: csub,
                COL_NOTE: cnote,
                COL_STATUS: STATUS_OK,
                COL_REASON: f"✔ {src_label}",
                COL_CONF: "",
                COL_METHOD: "category / تصنيف موثق",
                COL_REF: "",
                COL_SRC: source,
            })
            continue

        # صفوف المراجعة: تصنيف تقديري (قواعد/افتراضي) يحتاج تأكيدًا بشريًا
        ckey = csub.split(" - ")[0] if csub else None
        review_rows.append({
            COL_NAME: display_name,
            COL_CODE: resolve_code(src_code),
            COL_PRICE: price,
            COL_MAIN: cmain,
            COL_SUB: csub,
            COL_NOTE: cnote,
            COL_STATUS: STATUS_REVIEW,
            COL_REASON: build_reason(res, threshold, is_review=True)
            + (f" | تصنيف تقديري: {ckey}" if ckey else ""),
            COL_CONF: res["score"],
            COL_METHOD: res["method"],
            COL_REF: res["matched"],
            COL_SRC: source,
        })

    matched_df = pd.DataFrame(matched_rows, columns=ALL_COLS)
    review_df = pd.DataFrame(review_rows, columns=ALL_COLS)

    # الموثق أولًا مرتبًا بالتصنيف، ثم صفوف المراجعة (تُظلَّل حمراء) بالنهاية
    if not matched_df.empty:
        matched_df = matched_df.sort_values(
            by=[COL_MAIN, COL_SUB, COL_NAME]).reset_index(drop=True)
    if not review_df.empty:
        review_df = review_df.sort_values(
            by=[COL_MAIN, COL_SUB, COL_NAME]).reset_index(drop=True)
    df_all = pd.concat([matched_df, review_df], ignore_index=True)

    # تحقق نهائي: كل صف يجب أن يرتبط بتصنيف معتمد (فرعي + رئيسي = الأب)
    approved_labels = {c["label"] for c in cats.values()}
    n_bad = 0
    for df_chk in (matched_df, review_df):
        if df_chk.empty:
            continue
        bad = ~df_chk[COL_SUB].fillna("").isin(approved_labels)
        n_bad += int(bad.sum())
    if n_bad:
        print(f"⚠ تحذير: {n_bad} صفًا بتصنيف خارج القائمة المعتمدة — راجِعها!")
    else:
        print("✓ كل الصفوف مرتبطة بتصنيفات معتمدة")

    total = len(matched_df) + len(review_df)
    n_match = len(matched_df)
    n_review = len(review_df)
    summary = pd.DataFrame({
        "البند": ["إجمالي الخدمات الواردة", "جاهزة للرفع (Result)",
                  "— منها مطابقة سعرية بمرجع العقد",
                  "— منها جاهزة بتصنيف موثق (أودو/نوع المرفق)",
                  "تحتاج مراجعة (Need Review)", "نسبة الجاهزية",
                  "حد القبول (Threshold)", "محرك المطابقة",
                  "ملف التصنيفات المعتمدة", "صفوف بتصنيف غير معتمد"],
        "القيمة": [total, n_match,
                   n_match - n_cat_trusted[0], n_cat_trusted[0],
                   n_review,
                   f"{(n_match/total*100 if total else 0):.1f}%",
                   threshold, _FUZZ_ENGINE,
                   os.path.basename(categories_path), n_bad],
    })

    write_output(output, df_all, n_review, summary, cats)
    print(f"\n✓ تم الإنشاء: {output}")
    print(f"  - جاهزة: {n_match} (مطابقة عقد: {n_match - n_cat_trusted[0]}"
          f" + تصنيف موثق: {n_cat_trusted[0]})  |  تحتاج مراجعة تصنيف: {n_review}"
          f"  |  الإجمالي: {total}")
    return output


def main():
    p = argparse.ArgumentParser(
        description="مُرتِّب ومُطابِق قوائم الأسعار لنظام TPA")
    p.add_argument("--reference", required=True, help="ملف المرجع (الكتالوج الرسمي)")
    p.add_argument("--input", required=True, help="قائمة الأسعار الواردة (فوضوية)")
    p.add_argument("--output", required=True, help="ملف الإخراج المرتّب")
    p.add_argument("--synonyms", default=None, help="ملف قاموس المرادفات JSON (اختياري)")
    p.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD,
                   help=f"حد قبول المطابقة 0-100 (افتراضي {DEFAULT_THRESHOLD})")
    p.add_argument("--code-prefix", default="NEW",
                   help="بادئة الكود المولّد للخدمات بلا كود (افتراضي NEW)")
    p.add_argument("--categories", default=None,
                   help="ملف قائمة التصنيفات المعتمدة "
                        "(افتراضي: قائمة التصنيفات المعتمدة.xlsx بجانب السكربت)")
    p.add_argument("--hint", default=None, choices=GROUP_HINTS,
                   help="تلميح نوع المرفق: dental (أسنان) / optics (بصريات) "
                        "/ physio (علاج طبيعي)")
    args = p.parse_args()
    process(args.reference, args.input, args.output, args.synonyms,
            args.threshold, args.code_prefix, args.categories, args.hint)


if __name__ == "__main__":
    main()
