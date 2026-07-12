#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
  collect_approved.py  |  أداة إنتاج ملف الرفع النهائي
=============================================================================
بعد مراجعة ورقة «الخدمات Services» (تأكيد/تصحيح التصنيف في الصفوف
الحمراء) وحفظ الملف، شغّل:

    python collect_approved.py "نتيجة_الخدمات.xlsx"

تُنتج الأداة ملفًا جديدًا «..._للاعتماد_النهائي.xlsx» يحتوي ورقة واحدة
بالأعمدة الستة لملف الرفع فقط، جاهزة لنسخها ولصقها مباشرة في ملف
الـ price_list المعتمد.
=============================================================================
"""
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_SHEET = "الخدمات Services"      # الورقة الموحدة
STD_COLS = [
    "service_name / اسم الخدمة ★",
    "service_code / الكود",
    "contract_price / سعر العقد",
    "main_category / التصنيف الرئيسي",
    "sub_category / البند (التصنيف الفرعي)",
    "notes / ملاحظات",
]


def _first6(df):
    """يأخذ أول 6 أعمدة ويعيد تسميتها بأسماء ملف الاعتماد القياسية."""
    out = df.iloc[:, :6].copy()
    out.columns = STD_COLS
    return out


def collect(path):
    if not os.path.exists(path):
        sys.exit(f"الملف غير موجود: {path}")
    xl = pd.ExcelFile(path)

    if DATA_SHEET not in xl.sheet_names:
        sys.exit(f"لم أجد ورقة «{DATA_SHEET}» في الملف. تأكد أنه ملف نتيجة بالتنسيق الحالي.")

    # الورقة الموحدة «الخدمات» — كل الصفوف بعد المراجعة
    final = _first6(pd.read_excel(xl, sheet_name=DATA_SHEET, dtype=str))
    final = final.fillna("")
    print(f"• صفوف ورقة الخدمات الموحدة: {len(final)}")
    return _write_final(path, final)


def _write_final(path, final):
    """يكتب ملف الاعتماد النهائي بورقة واحدة منسّقة."""
    base = os.path.splitext(path)[0]
    out_path = f"{base}_للاعتماد_النهائي.xlsx"
    final.to_excel(out_path, sheet_name="النتيجة النهائية", index=False)

    wb = load_workbook(out_path)
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    hdr_fill = PatternFill("solid", fgColor="1F6E43")
    for c in range(1, len(STD_COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        letter = get_column_letter(c)
        maxlen = max([len(str(ws.cell(row=r, column=c).value or ""))
                      for r in range(1, min(ws.max_row, 200) + 1)] + [12])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 14), 50)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(STD_COLS) + 1):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
    wb.save(out_path)
    print(f"\n✓ تم الإنشاء: {out_path}")
    print("  افتحه، انسخ كل الجدول، والصقه في ملف الـ price_list المعتمد.")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit('الاستخدام: python collect_approved.py "ملف_النتيجة.xlsx"')
    collect(sys.argv[1])
